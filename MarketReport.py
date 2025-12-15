import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
from StockPrice import StockClient, read_watchlist

if __name__ == "__main__":
    client = StockClient()
    symbols = read_watchlist()
    data_list = []

    for sym in symbols:
        result = client.fetch_price(sym)
        if result:
            data_list.append(result)

    if not data_list:
        print("No data fetched. Exiting.")
        exit()

    df = pd.DataFrame(data_list)
    df["Change %"] = ((df["current_price"] - df["previous_close"]) / df["previous_close"]) * 100

    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"Market_Report_{today}.xlsx"
    df.to_excel(filename, index=False)
    print(f"Excel report saved as {filename}")

    wb = load_workbook(filename)
    ws = wb.active

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        change_cell = row[3]
        if change_cell.value > 0:
            for cell in row:
                cell.font = Font(color="008000")
        elif change_cell.value < 0:
            for cell in row:
                cell.font = Font(color="FF0000")

    wb.save(filename)
    print(f"Excel report formatted and saved: {filename}")